# loesche_mitarbeiter.py

import os
from openpyxl import load_workbook

from PyQt6.QtWidgets import (
    QWidget, QLabel, QLineEdit, QPushButton,
    QMessageBox, QVBoxLayout, QHBoxLayout, QSpacerItem, QSizePolicy
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont

from pathlib import Path
import sys

# Basisordner bestimmen (funktioniert im Script UND in der EXE)
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

EXCEL_PATH = BASE_DIR / "Masterliste_UTeam.xlsx"
# falls du noch andere Dateien brauchst:
# WORD_VORLAGEN_DIR = BASE_DIR / "vorlagen"


class LoescheMitarbeiter(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window

        # Hauptlayout mit Randabstand
        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(30)

        # Titel
        lbl = QLabel("Mitarbeiter löschen")
        lbl.setFont(QFont("Helvetica", 20, QFont.Weight.Bold))
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(lbl)

        # Mittiger Spacer oben
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        # Eingabefelder: in der Mitte, mit fester Breite
        row = QHBoxLayout()
        row.addStretch()
        self.le_vor = QLineEdit()
        self.le_vor.setPlaceholderText("Vorname")
        self.le_vor.setFixedWidth(200)
        row.addWidget(self.le_vor, alignment=Qt.AlignmentFlag.AlignCenter)
        row.addSpacing(20)
        self.le_nach = QLineEdit()
        self.le_nach.setPlaceholderText("Nachname")
        self.le_nach.setFixedWidth(200)
        row.addWidget(self.le_nach, alignment=Qt.AlignmentFlag.AlignCenter)
        row.addStretch()
        layout.addLayout(row)

        # Spacer zwischen Formular und Buttons
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        # Button-Leiste, rechtsbündig
        btn_row = QHBoxLayout()
        btn_row.addStretch()

        btn_del = QPushButton("Löschen")
        btn_del.setStyleSheet("""
            background-color: #D32F2F;
            color: white;
            border-radius: 5px;
            padding: 6px 18px;
        """)
        btn_del.clicked.connect(self.on_delete)
        btn_row.addWidget(btn_del)

        btn_cancel = QPushButton("Abbrechen")
        btn_cancel.setStyleSheet("""
            background-color: #9E9E9E;
            color: white;
            border-radius: 5px;
            padding: 6px 18px;
        """)
        btn_cancel.clicked.connect(self.on_cancel)
        btn_row.addWidget(btn_cancel)

        layout.addLayout(btn_row)

    def on_cancel(self):
        """Form zurücksetzen und zur Startseite."""
        self.le_vor.clear()
        self.le_nach.clear()
        self.main_window.zeige_startseite()

    def on_delete(self):
        """Löscht den Eintrag in der Masterlist-Excel."""
        vor = self.le_vor.text().strip()
        nach = self.le_nach.text().strip()
        if not vor or not nach:
            QMessageBox.warning(self, "Fehler", "Bitte Vorname und Nachname eingeben.")
            return

        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb["Masterlist"]

            gefunden = False
            for row in ws.iter_rows(min_row=2, values_only=False):
                if str(row[0].value).strip() == vor and str(row[1].value).strip() == nach:
                    ws.delete_rows(row[0].row, 1)
                    gefunden = True
                    break

            if not gefunden:
                QMessageBox.information(self, "Nicht gefunden",
                                        "Kein Mitarbeiter mit diesem Namen in der Masterlist.")
                return

            wb.save(EXCEL_PATH)
        except Exception as e:
            QMessageBox.critical(self, "Excel-Fehler", f"Konnte nicht löschen:\n{e}")
            return

        QMessageBox.information(self, "Erfolg",
                                f"{vor} {nach} wurde aus der Masterlist gelöscht.")
        self.le_vor.clear()
        self.le_nach.clear()
        self.main_window.zeige_startseite()
