# excel_utils.py

from openpyxl import load_workbook

def add_employee_to_excel(excel_path: str, master_data: dict, rotplan_data: dict):
    """
    Fügt einen neuen Mitarbeiter in die Excel-Datei ein.
    - master_data: Werte für Masterlist-Blatt (Vorname, Nachname, Geburtsdatum, Aktueller Einsatz, ...)
    - rotplan_data: Werte für Rotationsplan-Blatt (Vorname, Nachname, ggf. erste Station)
    """
    wb = load_workbook(excel_path)

    # 1) Masterlist
    ws_master = wb["Masterlist"]
    neue_zeile_master = [
        master_data.get("Vorname", ""),
        master_data.get("Nachname", ""),
        master_data.get("Geburtsdatum", ""),
        master_data.get("Aktueller Einsatz", ""),
        # … hier je nach Spaltenreihenfolge weitere Felder ergänzen …
    ]
    ws_master.append(neue_zeile_master)

    # 2) Rotationsplan
    ws_rot = wb["Rotationsplan"]
    name = f"{rotplan_data.get('Vorname','')} {rotplan_data.get('Nachname','')}"
    neue_zeile_rot = [name]
    erste_station = rotplan_data.get("Aktueller Einsatz", "")
    neue_zeile_rot.append(erste_station)
    # Restliche Spalten (Station 2–8) bleiben leer
    for _ in range(ws_rot.max_column - len(neue_zeile_rot)):
        neue_zeile_rot.append("")
    ws_rot.append(neue_zeile_rot)

    # Speichern
    wb.save(excel_path)
