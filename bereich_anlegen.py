# bereich_anlegen.py

import os
from PyQt6.QtWidgets import (
    QWidget, QLabel, QComboBox, QLineEdit, QPushButton,
    QHBoxLayout, QVBoxLayout, QMessageBox, QFrame
)
from PyQt6.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from pathlib import Path
import sys

EXCEL_PATH = r"C:\Users\ASYAKKA\Mercedes-Benz (corpdir.onmicrosoft.com)\DWT_UTeam Werk 10 - General\08_Rotation UTeam\Projekt_UTeam_Digitalisierung\Masterliste_UTeam.xlsx"
OUTPUT_DIR  = r"C:\Users\ASYAKKA\Mercedes-Benz (corpdir.onmicrosoft.com)\DWT_UTeam Werk 10 - General\Vorstellung_Asya_Test"


class BereichAnlegen(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window

        # Layout
        v = QVBoxLayout(self)
        v.setContentsMargins(30,30,30,30)
        v.setSpacing(20)

        lbl = QLabel("Neuen Bereich anlegen")
        lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl.setStyleSheet("font-size:18pt; font-weight:bold;")
        v.addWidget(lbl)

        # Trennlinie
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        v.addWidget(line)

        # Auswahl und Eingabe
        from_layout = QHBoxLayout()
        v.addLayout(from_layout)

        # 1) Drop‑down für Kategorie
        from_layout.addWidget(QLabel("Bereich:"))
        self.cb_kategorie = QComboBox()
        self.cb_kategorie.addItems(["eATS","Gießerei","Montage","Logistik","Qualität","Fertigung"])
        from_layout.addWidget(self.cb_kategorie)

        # 2) Feld für Namen (z.B. „Batterie Montage“)
        from_layout.addSpacing(30)
        from_layout.addWidget(QLabel("Name:"))
        self.le_name = QLineEdit()
        self.le_name.setPlaceholderText("z.B. Batterie Montage")
        from_layout.addWidget(self.le_name, stretch=1)

        # 3) Kostenstelle mit Prefix „010-“
        from_layout.addSpacing(30)
        from_layout.addWidget(QLabel("Kostenstelle:"))
        self.le_kst = QLineEdit("010-")
        self.le_kst.setFixedWidth(120)
        from_layout.addWidget(self.le_kst)

        v.addStretch()

        # Buttons
        btns = QHBoxLayout()
        btns.addStretch()

        self.btn_save = QPushButton("Speichern")
        self.btn_save.clicked.connect(self.on_save)
        btns.addWidget(self.btn_save)

        self.btn_cancel = QPushButton("Abbrechen")
        self.btn_cancel.clicked.connect(self.on_cancel)
        btns.addWidget(self.btn_cancel)

        v.addLayout(btns)

    def on_cancel(self):
        """Zurück zur Startseite."""
        self.main_window.zeige_startseite()

    def on_save(self):
        """Neuen Bereich in Excel eintragen (unterhalb der Kategorie)."""
        kategorie = self.cb_kategorie.currentText().strip()
        name      = self.le_name.text().strip()
        kst       = self.le_kst.text().strip()

        if not name or not kst.startswith("010-"):
            QMessageBox.warning(self, "Fehler",
                "Bitte einen gültigen Namen eingeben und die Kostenstelle\n"
                "mit dem Prefix 010- beginnen lassen.")
            return

        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb["Bereich"]

            # 1) Merge‑Blöcke in Zeile1 auflösen und Titel kopieren
            for merged in list(ws.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_boundaries(str(merged))
                if min_row == 1:
                    text = ws.cell(row=1, column=min_col).value
                    for col in range(min_col, max_col+1):
                        ws.cell(row=1, column=col, value=text)
                    ws.unmerge_cells(str(merged))

            # 2) Indexiere alle Header aus Zeile1
            headers = {}
            for col in range(1, ws.max_column+1):
                val = ws.cell(row=1, column=col).value
                if isinstance(val, str):
                    headers[val.strip().lower()] = col

            key = kategorie.lower()
            if key not in headers:
                raise KeyError(f"Spalte '{kategorie}' nicht gefunden.")

            col_idx = headers[key]
            # finde die erste leere Zeile in dieser Spalte (unter Zeile1)
            row_idx = 2
            while ws.cell(row=row_idx, column=col_idx).value not in (None, ""):
                row_idx += 1

            # 3) Werte eintragen
            ws.cell(row=row_idx, column=col_idx, value=name)
            # in der Nachbarspalte rechts (col_idx+1) die Kostenstelle
            ws.cell(row=row_idx, column=col_idx+1, value=kst)

            wb.save(EXCEL_PATH)

        except KeyError as ke:
            QMessageBox.critical(self, "Excel-Fehler", str(ke))
            return
        except Exception as e:
            QMessageBox.critical(self, "Fehler beim Speichern", str(e))
            return

        QMessageBox.information(self, "Erfolg",
            f"Bereich '{name}' unter '{kategorie}' hinzugefügt.")
        # Felder zurücksetzen
        self.le_name.clear()
        self.le_kst.setText("010-")
        self.main_window.zeige_startseite()
